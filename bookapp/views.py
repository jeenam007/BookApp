from django.shortcuts import render, get_object_or_404, redirect
from .models import Book,Account
from .forms import BookForm,UploadFileForm,LoginForm,UserRegistrationForm
from django.contrib import messages
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.db import IntegrityError
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required
from io import BytesIO

@login_required
def home_view(request):
    return render(request,'index.html')

def login_view(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                print('login success')
                
                return redirect('home')
            else:
                print('login failed')
                messages.error(request, "Invalid username or password")
                
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form':form})

def sign_out(request):
     logout(request)
     return redirect('userlogin')

def registration(request):
    
    if request.method =="POST":
       form=UserRegistrationForm(request.POST)
       if form.is_valid():
          form.save()
          messages.success(request, "User registered successfully!")
          return redirect("userlist")
       else:
          messages.error(request,"Registration failed.")
    else:
        form=UserRegistrationForm()
    return render(request,'registration.html',{'form':form})

def userlist(request):
        users=Account.objects.all()
        return render(request,'userlist.html',{'users':users})

# Create
def book_create(request):
    form = BookForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('book_list')
    return render(request, 'book_form.html', {'form': form})

# Read
def book_list(request):
    books = Book.objects.all()
    return render(request, 'book_list.html', {'books': books})

# Update
def book_update(request, pk):
    book = get_object_or_404(Book, pk=pk)
    form = BookForm(request.POST or None, instance=book)

    # instance_edit=Book.objects.get(pk=pk)
    # if request.POST:
    #     form=BookForm(request.POST,instance=instance_edit)
    if form.is_valid():
        form.save()
        return redirect('book_list')
    return render(request, 'book_form.html', {'form': form})

# Delete
def book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk)
    book.delete()
    return redirect('book_list')
    rem_list=Book.objects.all()
    return render(request, 'book_list.html', {'rem_list': rem_list})
#Detail
def book_detail(request,pk):
     book=get_object_or_404(Book, pk=pk)
     return render(request,'book_detail.html',{'book':book})

def is_empty(value):
    return pd.isna(value) or str(value).strip() == ''

def upload_excel(request):
    form = UploadFileForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        excel_file = request.FILES['file']
        wb = load_workbook(excel_file)
        sheet = wb.active
           
        created = 0
        error_rows = []

        # Header map
        headers = [cell.value.strip().lower().replace('*', '') if cell.value else '' for cell in sheet[1]]

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell for cell in row if cell is not None and str(cell).strip() != ''):
                continue
            title,author,price,gmail,published_date=row[0],row[1],row[2],row[3],row[4]
            error_message = ""


           
            # Map row to headers
            # row_data = dict(zip(headers, row))
            # title = str(row_data.get('title', '')).strip()
            # author = str(row_data.get('author', '')).strip()
            # price = row_data.get('price')
            # gmail = str(row_data.get('gmail', '')).strip()
            # published_date = row_data.get('published_date')

            if not all([title,author,price,gmail]):
                error_message="Missing required fields."
            
            # # Validation
            # missing_fields = []
            # if not title:
            #     missing_fields.append('title')
            # if not author:
            #     missing_fields.append('author')
            # if price is None or str(price).strip() == '':
            #     missing_fields.append('price')
            # if not published_date:
            #     missing_fields.append('published_date')

            if error_message:
                error_rows.append((
                    idx, title, author, price, gmail, published_date,error_message))
                continue

            try:
                Book.objects.create(
                    title=title,
                    author=author,
                    price=price,
                    gmail=gmail,
                    published_date=published_date
                )
                created += 1
            except IntegrityError as e:
                error_rows.append((
                    idx, title, author, price, gmail, published_date,
                    f"Database error: {str(e)}"
                ))

        # Export error rows if any
        if error_rows:
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.title = "Errors"
            error_ws.append(['Row', 'Title', 'Author', 'Price', 'Gmail', 'Published Date', 'Error'])

            for error_row in error_rows:
                error_ws.append(error_row)

            output = BytesIO()
            error_wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=book_upload_errors.xlsx'
            return response

        messages.success(request, f"{created} books uploaded successfully.")
        return redirect('book_list')

    return render(request, 'upload_excel.html', {'form': form})


    

def export_books_excel(request):
    # Get the data from the database
    books = Book.objects.all().values('title', 'author', 'price','gmail', 'published_date')  
    df = pd.DataFrame(books)
    df.columns = ['Book Title', 'Author Name', 'Price ','Email', 'Publication Date']

    # Create a response to send as an Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=books.xlsx'

    # Create a workbook and add a sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Books'

    # Convert the DataFrame to rows and write to the sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Get the headers row (first row)
    header_row = ws[1]

    # Apply bold font to headers
    for cell in header_row:
        cell.font = Font(bold=True)

    # Apply borders to all cells
    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    # Adjust the column widths to fit the contents
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to the response
    wb.save(response)

    return response